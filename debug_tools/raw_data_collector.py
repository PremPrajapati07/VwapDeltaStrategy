"""
raw_data_collector.py
=====================
Fetches 1-minute raw OHLCV + OI + Greeks + IV for all 22 Nifty option
instruments (11 strikes × CE + PE) on a given trading day and saves
them into a single day-wise Excel file.

File naming  : NIFTY_24Mar2026.xlsx
Tab naming   : "24000 CE", "24000 PE", "24050 CE", … (22 tabs)
Columns/tab  : Time | Open | High | Low | Close | Volume |
               OI   | OI Change | IV% | Delta | Gamma | Theta | Vega

Greeks are computed via Black-Scholes for historical data.
For live intraday, Greeks are taken directly from Kite quotes if available,
else computed on the fly.

Usage
-----
  # Fetch today (live / post-market backfill):
  python raw_data_collector.py

  # Fetch a specific past date:
  python raw_data_collector.py --date 2026-03-20

  # Fetch a date range (all Thursdays):
  python raw_data_collector.py --start 2026-01-01 --end 2026-03-20

  # Dry-run: generate a sample file with synthetic data (no Kite needed):
  python raw_data_collector.py --demo
"""

import os
import math
import datetime as dt
import argparse
import logging
import time
import random
from typing import Optional

import numpy as np
from scipy.stats import norm
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

OUTPUT_DIR   = "raw_data"
LOG_DIR      = "logs"
STRIKE_STEP  = 50
N_EACH_SIDE  = 5          # 5 strikes above + ATM + 5 below = 11 strikes
MARKET_OPEN  = dt.time(9, 15)
MARKET_CLOSE = dt.time(15, 30)
RISK_FREE    = 0.068       # ~6.8% annualised (approximate Indian risk-free)

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(LOG_DIR,    exist_ok=True)

logging.basicConfig(
    filename=os.path.join(LOG_DIR, "raw_collector.log"),
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s"
)
log = logging.getLogger(__name__)


# ── Black-Scholes Greeks ─────────────────────────────────────

def bs_greeks(S: float, K: float, T: float, r: float, sigma: float,
              option_type: str) -> dict:
    """
    Compute Black-Scholes price + Greeks.
    S     : spot price
    K     : strike
    T     : time to expiry in years  (e.g. 5/365)
    r     : risk-free rate (annual)
    sigma : implied volatility (annual, decimal)
    option_type : 'CE' or 'PE'
    Returns dict with keys: price, delta, gamma, theta, vega, iv
    """
    if T <= 0 or sigma <= 0 or S <= 0:
        return dict(price=0, delta=0, gamma=0, theta=0, vega=0, iv=sigma*100)

    d1 = (math.log(S / K) + (r + 0.5 * sigma**2) * T) / (sigma * math.sqrt(T))
    d2 = d1 - sigma * math.sqrt(T)

    if option_type == 'CE':
        price = S * norm.cdf(d1) - K * math.exp(-r * T) * norm.cdf(d2)
        delta = norm.cdf(d1)
    else:
        price = K * math.exp(-r * T) * norm.cdf(-d2) - S * norm.cdf(-d1)
        delta = norm.cdf(d1) - 1

    gamma = norm.pdf(d1) / (S * sigma * math.sqrt(T))
    vega  = S * norm.pdf(d1) * math.sqrt(T) / 100      # per 1% IV move
    theta = (-(S * norm.pdf(d1) * sigma) / (2 * math.sqrt(T))
             - r * K * math.exp(-r * T) * (norm.cdf(d2) if option_type == 'CE'
                                            else norm.cdf(-d2))) / 365

    return dict(
        price=round(price, 2),
        delta=round(delta, 4),
        gamma=round(gamma, 6),
        theta=round(theta, 4),
        vega =round(vega,  4),
        iv   =round(sigma * 100, 2)
    )


def implied_vol(market_price: float, S: float, K: float, T: float,
                r: float, option_type: str,
                tol: float = 1e-5, max_iter: int = 100) -> float:
    """Newton-Raphson IV solver. Returns annualised IV (decimal)."""
    if T <= 0 or market_price <= 0:
        return 0.0
    sigma = 0.20    # initial guess
    for _ in range(max_iter):
        g = bs_greeks(S, K, T, r, sigma, option_type)
        price_diff = g['price'] - market_price
        if abs(price_diff) < tol:
            break
        vega_raw = S * norm.pdf(
            (math.log(S/K) + (r + 0.5*sigma**2)*T) / (sigma*math.sqrt(T))
        ) * math.sqrt(T)
        if vega_raw < 1e-10:
            break
        sigma -= price_diff / vega_raw
        sigma = max(0.001, min(sigma, 5.0))
    return sigma


# ── Kite Integration (graceful import) ──────────────────────

def _try_import_kite():
    try:
        from kiteconnect import KiteConnect
        return KiteConnect
    except ImportError:
        return None


def get_kite(api_key: str, access_token: str):
    KiteConnect = _try_import_kite()
    if KiteConnect is None:
        raise ImportError("kiteconnect not installed. Run: pip install kiteconnect")
    kite = KiteConnect(api_key=api_key)
    kite.set_access_token(access_token)
    return kite


def load_access_token() -> tuple[str, str]:
    token_file = ".access_token"
    keys_file  = ".kite_keys"
    if not os.path.exists(token_file) or not os.path.exists(keys_file):
        raise FileNotFoundError(
            "Missing .access_token or .kite_keys file.\n"
            "Run: python kite_login.py  to authenticate first."
        )
    with open(token_file) as f:
        access_token = f.read().strip()
    with open(keys_file) as f:
        api_key = f.read().strip().split('\n')[0]
    return api_key, access_token


def get_nifty_spot_at(kite, trade_date: dt.date) -> float:
    """Fetch Nifty spot close at 9:20 for given date."""
    start = dt.datetime.combine(trade_date, dt.time(9, 19))
    end   = dt.datetime.combine(trade_date, dt.time(9, 21))
    candles = kite.historical_data(256265, start, end, "minute")
    if not candles:
        raise ValueError(f"No Nifty spot data for {trade_date}")
    return candles[-1]['close']


def get_next_thursday(from_date: dt.date) -> dt.date:
    days = (3 - from_date.weekday()) % 7
    return from_date + dt.timedelta(days=days if days > 0 else 7)


def get_instruments(kite, trade_date: dt.date, spot: float, expiry: dt.date):
    """Return list of 22 instrument dicts {strike, type, token, symbol}."""
    atm = round(spot / STRIKE_STEP) * STRIKE_STEP
    strikes = [atm + i * STRIKE_STEP
               for i in range(-N_EACH_SIDE, N_EACH_SIDE + 1)]

    instr_df = kite.instruments("NFO")
    import pandas as pd
    df = pd.DataFrame(instr_df)
    expiry_str = expiry.strftime("%Y-%m-%d")
    nifty = df[
        (df["name"] == "NIFTY") &
        (df["expiry"].astype(str) == expiry_str) &
        (df["instrument_type"].isin(["CE", "PE"]))
    ]

    result = []
    for strike in strikes:
        for opt_type in ["CE", "PE"]:
            row = nifty[
                (nifty["strike"] == strike) &
                (nifty["instrument_type"] == opt_type)
            ]
            if row.empty:
                log.warning(f"No instrument: NIFTY {strike} {opt_type} {expiry_str}")
                continue
            result.append({
                "strike":   int(strike),
                "type":     opt_type,
                "token":    int(row.iloc[0]["instrument_token"]),
                "symbol":   row.iloc[0]["tradingsymbol"],
                "atm":      atm,
                "expiry":   expiry,
            })
    return result, atm


def fetch_minute_candles(kite, token: int, trade_date: dt.date) -> list[dict]:
    """Fetch 1-min OHLCV from 9:15 to 15:30."""
    start = dt.datetime.combine(trade_date, MARKET_OPEN)
    end   = dt.datetime.combine(trade_date, MARKET_CLOSE)
    try:
        return kite.historical_data(token, start, end, "minute")
    except Exception as e:
        log.error(f"Fetch failed for token {token}: {e}")
        return []


# ── Demo / Synthetic Data ────────────────────────────────────

def generate_synthetic_candles(strike: int, opt_type: str,
                                spot: float, expiry: dt.date,
                                trade_date: dt.date) -> list[dict]:
    """Generate realistic synthetic 1-min data for demo mode."""
    candles = []
    current  = dt.datetime.combine(trade_date, MARKET_OPEN)
    end_dt   = dt.datetime.combine(trade_date, MARKET_CLOSE)
    days_to_exp = max((expiry - trade_date).days, 0.001)
    T = days_to_exp / 365.0
    base_iv = 0.16 + random.uniform(-0.02, 0.03)
    base_spot = spot

    prev_oi = random.randint(800000, 2500000)
    prev_close = None

    while current <= end_dt:
        elapsed_min = (current - dt.datetime.combine(trade_date, MARKET_OPEN)).seconds // 60
        spot_drift  = base_spot * (1 + random.gauss(0, 0.0008))
        base_spot   = spot_drift
        iv          = max(0.08, base_iv + random.gauss(0, 0.002))
        T_now       = max((end_dt - current).seconds / (365 * 24 * 3600), 0.0001)

        g = bs_greeks(base_spot, strike, T_now, RISK_FREE, iv, opt_type)
        close  = max(0.05, g['price'] + random.gauss(0, 0.5))
        open_  = prev_close if prev_close else close * (1 + random.uniform(-0.005, 0.005))
        high   = max(open_, close) * (1 + random.uniform(0, 0.008))
        low    = min(open_, close) * (1 - random.uniform(0, 0.008))
        volume = random.randint(500, 15000)
        oi_chg = random.randint(-5000, 8000)
        oi     = max(0, prev_oi + oi_chg)

        iv_now = implied_vol(close, base_spot, strike, T_now, RISK_FREE, opt_type)
        g2     = bs_greeks(base_spot, strike, T_now, RISK_FREE, iv_now, opt_type)

        candles.append({
            "date":     current,
            "open":     round(open_, 2),
            "high":     round(high, 2),
            "low":      round(low, 2),
            "close":    round(close, 2),
            "volume":   volume,
            "oi":       oi,
            "oi_change":oi_chg,
            "iv":       round(iv_now * 100, 2),
            "delta":    g2['delta'],
            "gamma":    g2['gamma'],
            "theta":    g2['theta'],
            "vega":     g2['vega'],
        })
        prev_close = close
        prev_oi    = oi
        current   += dt.timedelta(minutes=1)

    return candles


def fetch_and_compute(kite, instr: dict, trade_date: dt.date,
                      spot: float, demo: bool = False) -> list[dict]:
    """
    Fetch raw OHLCV from Kite, then compute IV + Greeks per minute.
    Falls back to demo data if demo=True or fetch fails.
    """
    if demo:
        return generate_synthetic_candles(
            instr['strike'], instr['type'], spot,
            instr['expiry'], trade_date
        )

    raw = fetch_minute_candles(kite, instr['token'], trade_date)
    if not raw:
        return []

    result = []
    prev_oi = raw[0].get('oi', 0) if raw else 0
    expiry  = instr['expiry']

    for candle in raw:
        ts      = candle['date']
        close   = candle['close']
        oi      = candle.get('oi', prev_oi)
        oi_chg  = oi - prev_oi
        T_now   = max((dt.datetime.combine(expiry, dt.time(15, 30)) - ts
                       ).total_seconds() / (365 * 24 * 3600), 1e-6)

        iv_val  = implied_vol(close, spot, instr['strike'], T_now,
                              RISK_FREE, instr['type'])
        g       = bs_greeks(spot, instr['strike'], T_now,
                            RISK_FREE, iv_val, instr['type'])

        result.append({
            "date":      ts,
            "open":      candle.get('open',  close),
            "high":      candle.get('high',  close),
            "low":       candle.get('low',   close),
            "close":     close,
            "volume":    candle.get('volume', 0),
            "oi":        oi,
            "oi_change": oi_chg,
            "iv":        g['iv'],
            "delta":     g['delta'],
            "gamma":     g['gamma'],
            "theta":     g['theta'],
            "vega":      g['vega'],
        })
        prev_oi = oi

    return result


# ── Excel Builder ────────────────────────────────────────────

COLUMNS = [
    ("Time",      12, None),
    ("Open",      10, '#,##0.00'),
    ("High",      10, '#,##0.00'),
    ("Low",       10, '#,##0.00'),
    ("Close",     10, '#,##0.00'),
    ("Volume",    12, '#,##0'),
    ("OI",        14, '#,##0'),
    ("OI Change", 12, '+#,##0;-#,##0;-'),
    ("IV %",      9,  '0.00"%"'),
    ("Delta",     9,  '0.0000'),
    ("Gamma",     9,  '0.000000'),
    ("Theta",     9,  '0.0000'),
    ("Vega",      9,  '0.0000'),
]

GROUP_SPANS = [
    ("OHLC",      1, 4,  "1F77B4"),   # blue   col B-E
    ("Volume",    5, 5,  "2CA02C"),   # green  col F
    ("OI",        6, 7,  "FF7F0E"),   # orange col G-H
    ("Greeks + IV", 8, 12, "9467BD"), # purple col I-M
]

HDR_BG_GROUP = {
    "OHLC":        "D9E8F7",
    "Volume":      "D5EDCF",
    "OI":          "FCE5C5",
    "Greeks + IV": "E8E4F9",
}
HDR_FONT_GROUP = {
    "OHLC":        "185FA5",
    "Volume":      "3B6D11",
    "OI":          "854F0B",
    "Greeks + IV": "534AB7",
}

CE_TAB_COLOR = "4472C4"   # blue
PE_TAB_COLOR = "C0504D"   # red


def _border(style='thin'):
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def build_sheet(ws, instr: dict, candles: list[dict], trade_date: dt.date):
    """Write one instrument's data into worksheet ws."""
    opt_type = instr['type']
    strike   = instr['strike']
    atm      = instr['atm']
    symbol   = instr.get('symbol', f"NIFTY{strike}{opt_type}")

    ws.sheet_properties.tabColor = CE_TAB_COLOR if opt_type == 'CE' else PE_TAB_COLOR

    # ── Row 1: instrument header ──────────────────────────────
    ws.merge_cells("A1:M1")
    ws["A1"] = f"{symbol}  |  Strike: {strike}  |  {'ATM' if strike == atm else f'{abs(strike-atm)} pts from ATM'}  |  {trade_date.strftime('%d %b %Y')}"
    ws["A1"].font      = Font(bold=True, size=12, color="FFFFFF",
                              name="Arial")
    ws["A1"].fill      = _fill(CE_TAB_COLOR if opt_type == 'CE' else PE_TAB_COLOR)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Row 2: group labels ───────────────────────────────────
    ws.row_dimensions[2].height = 16
    for group_name, col_start, col_end, _ in GROUP_SPANS:
        start_letter = get_column_letter(col_start + 1)   # +1 for Time col A
        end_letter   = get_column_letter(col_end + 1)
        ws.merge_cells(f"{start_letter}2:{end_letter}2")
        cell = ws[f"{start_letter}2"]
        cell.value     = group_name
        cell.font      = Font(bold=True, size=9, color=HDR_FONT_GROUP[group_name], name="Arial")
        cell.fill      = _fill(HDR_BG_GROUP[group_name])
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Time cell row 2
    ws["A2"].fill      = _fill("F0F0F0")
    ws["A2"].border    = _border()

    # ── Row 3: column headers ─────────────────────────────────
    ws.row_dimensions[3].height = 18
    for col_idx, (col_name, col_width, _) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        cell   = ws[f"{letter}3"]
        cell.value     = col_name
        cell.font      = Font(bold=True, size=10, name="Arial",
                              color="333333")
        cell.fill      = _fill("F5F5F5")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border()
        ws.column_dimensions[letter].width = col_width

    # ── Data rows ─────────────────────────────────────────────
    for row_idx, candle in enumerate(candles, start=4):
        row_data = [
            candle['date'].strftime("%H:%M"),
            candle['open'],
            candle['high'],
            candle['low'],
            candle['close'],
            candle['volume'],
            candle['oi'],
            candle['oi_change'],
            candle['iv'],
            candle['delta'],
            candle['gamma'],
            candle['theta'],
            candle['vega'],
        ]
        bg = "FFFFFF" if row_idx % 2 == 0 else "FAFAFA"

        for col_idx, (value, (_, _, num_fmt)) in enumerate(
                zip(row_data, COLUMNS), start=1):
            letter = get_column_letter(col_idx)
            cell   = ws[f"{letter}{row_idx}"]
            cell.value     = value
            cell.font      = Font(size=10, name="Arial")
            cell.fill      = _fill(bg)
            cell.border    = _border()
            cell.alignment = Alignment(
                horizontal="left" if col_idx == 1 else "right",
                vertical="center"
            )
            if num_fmt:
                cell.number_format = num_fmt

            # Colour OI Change positive green / negative red
            if col_idx == 8 and isinstance(value, (int, float)):
                if value > 0:
                    cell.font = Font(size=10, name="Arial", color="2E7D32")
                elif value < 0:
                    cell.font = Font(size=10, name="Arial", color="C62828")

    # ── Freeze panes below header rows ───────────────────────
    ws.freeze_panes = "A4"


def build_excel(trade_date: dt.date, instruments: list[dict],
                all_candles: dict) -> str:
    """
    Build the full day Excel file with 22 tabs.
    all_candles: dict keyed by (strike, type) → list of candle dicts
    Returns output file path.
    """
    wb = Workbook()
    wb.remove(wb.active)    # remove default blank sheet

    # Order tabs: 24000 CE, 24000 PE, 24050 CE, 24050 PE, …
    strikes_sorted = sorted(set(i['strike'] for i in instruments))
    for strike in strikes_sorted:
        for opt_type in ['CE', 'PE']:
            key   = (strike, opt_type)
            instr = next((i for i in instruments
                          if i['strike'] == strike and i['type'] == opt_type), None)
            if instr is None:
                continue
            candles = all_candles.get(key, [])
            tab_name = f"{strike} {opt_type}"
            ws = wb.create_sheet(title=tab_name)
            build_sheet(ws, instr, candles, trade_date)
            log.info(f"Built sheet: {tab_name} ({len(candles)} rows)")

    date_str  = trade_date.strftime("%d%b%Y")    # e.g. 24Mar2026
    file_name = f"NIFTY_{date_str}.xlsx"
    file_path = os.path.join(OUTPUT_DIR, file_name)
    wb.save(file_path)
    log.info(f"Saved: {file_path}")
    return file_path


# ── Main Orchestrator ────────────────────────────────────────

def collect_day(trade_date: dt.date, demo: bool = False,
                kite=None, spot_override: Optional[float] = None) -> str:
    """
    Full pipeline for one trading day:
    1. Get spot at 9:20 → derive 11 strikes
    2. Fetch 1-min candles for all 22 instruments
    3. Compute Greeks per minute
    4. Write Excel file
    Returns path to saved file.
    """
    log.info(f"=== Collecting {trade_date} (demo={demo}) ===")
    print(f"\n{'='*55}")
    print(f"  Nifty Raw Data Collector — {trade_date.strftime('%d %b %Y')}")
    print(f"  Mode: {'DEMO (synthetic data)' if demo else 'LIVE (Kite API)'}")
    print(f"{'='*55}")

    expiry = get_next_thursday(trade_date)

    # ── Step 1: Spot price ──
    if demo:
        spot = spot_override or 24000.0
        print(f"  Spot (demo): {spot:.0f}  |  Expiry: {expiry}")
    else:
        spot = get_nifty_spot_at(kite, trade_date)
        print(f"  Spot @ 9:20: {spot:.0f}  |  Expiry: {expiry}")

    # ── Step 2: Build instrument list ──
    if demo:
        atm = round(spot / STRIKE_STEP) * STRIKE_STEP
        strikes = [atm + i * STRIKE_STEP
                   for i in range(-N_EACH_SIDE, N_EACH_SIDE + 1)]
        instruments = [
            {"strike": s, "type": t, "token": 0,
             "symbol": f"NIFTY{s}{t}", "atm": atm, "expiry": expiry}
            for s in strikes for t in ["CE", "PE"]
        ]
    else:
        instruments, atm = get_instruments(kite, trade_date, spot, expiry)

    print(f"  ATM: {atm}  |  Strikes: {instruments[0]['strike']} -> {instruments[-1]['strike']}")
    print(f"  Instruments: {len(instruments)} (22 expected)")
    print()

    # ── Step 3: Fetch candles + compute Greeks ──
    all_candles = {}
    total = len(instruments)
    for idx, instr in enumerate(instruments, 1):
        key  = (instr['strike'], instr['type'])
        label = f"{instr['strike']} {instr['type']}"
        print(f"  [{idx:>2}/{total}] Fetching {label} …", end=" ", flush=True)
        candles = fetch_and_compute(kite, instr, trade_date, spot, demo=demo)
        all_candles[key] = candles
        print(f"{len(candles)} rows")
        if not demo:
            time.sleep(0.35)    # Kite rate limit

    # ── Step 4: Build Excel ──
    print(f"\n  Building Excel file …")
    file_path = build_excel(trade_date, instruments, all_candles)
    size_kb   = os.path.getsize(file_path) // 1024
    print(f"  Saved -> {file_path}  ({size_kb} KB)")
    print(f"  Tabs : 22  |  Rows/tab: ~{len(next(iter(all_candles.values())))} minutes\n")
    return file_path


def collect_range(start_date: dt.date, end_date: dt.date,
                  demo: bool = False, kite=None, thursdays_only: bool = True):
    """Collect data for all dates (or Thursdays only) in range."""
    d = start_date
    files = []
    while d <= end_date:
        if not thursdays_only or d.weekday() == 3:
            try:
                fp = collect_day(d, demo=demo, kite=kite)
                files.append(fp)
            except Exception as e:
                log.error(f"Failed {d}: {e}")
                print(f"  ERROR on {d}: {e}")
        d += dt.timedelta(days=1)
    print(f"\nDone. {len(files)} files saved to ./{OUTPUT_DIR}/")
    return files


# ── CLI ──────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Nifty Options Raw Data Collector → Day-wise Excel"
    )
    parser.add_argument("--date",  help="Single date YYYY-MM-DD")
    parser.add_argument("--start", help="Range start YYYY-MM-DD")
    parser.add_argument("--end",   help="Range end YYYY-MM-DD")
    parser.add_argument("--demo",  action="store_true",
                        help="Run with synthetic data (no Kite needed)")
    parser.add_argument("--spot",  type=float, default=24000,
                        help="Spot price for demo mode (default: 24000)")
    parser.add_argument("--all-days", action="store_true",
                        help="Collect every trading day, not just Thursdays")
    args = parser.parse_args()

    kite = None
    if not args.demo:
        api_key, access_token = load_access_token()
        kite = get_kite(api_key, access_token)
        print("Kite session loaded.")

    if args.date:
        collect_day(dt.date.fromisoformat(args.date),
                    demo=args.demo, kite=kite,
                    spot_override=args.spot)

    elif args.start and args.end:
        collect_range(
            dt.date.fromisoformat(args.start),
            dt.date.fromisoformat(args.end),
            demo=args.demo, kite=kite,
            thursdays_only=not args.all_days
        )

    else:
        # Default: today
        collect_day(dt.date.today(), demo=args.demo, kite=kite,
                    spot_override=args.spot)


if __name__ == "__main__":
    main()
